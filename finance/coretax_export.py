from __future__ import annotations

import json
import math
import re
import sqlite3
from dataclasses import asdict, dataclass, is_dataclass
from pathlib import Path
from typing import Any

import openpyxl

from finance.config import CoretaxConfig, get_coretax_config, load_config
from finance.db import open_db


@dataclass
class CoretaxRowTrace:
    xlsx_row: int
    raw_keterangan: str
    normalized: str
    parsed: dict | None
    matched: dict | None
    value_f: float | None
    value_g: float | None
    status: str
    warnings: list[str]


@dataclass
class CoretaxResult:
    snapshot_date: str
    filled_count: int
    skipped_out_of_scope: int
    unmatched_count: int
    aggregated_count: int
    currency_warning_count: int
    rows: list[CoretaxRowTrace]
    unused_pwm_rows: list[dict]
    output_path: Path | None


@dataclass(frozen=True)
class RowKey:
    kind: str
    institution: str | None
    account: str | None
    owner: str | None
    asset_class: str | None = None
    broker_present: bool = False
    broker_token: str | None = None
    idd_sid: str | None = None


class CoretaxTemplateError(ValueError):
    pass


_cfg = load_config()
_CORETAX_CONFIG = get_coretax_config(_cfg)
_OUT_OF_SCOPE_CODES = {"038", "042", "043", "051", "061"}
_INVESTMENT_CODE_TO_CLASS = {
    "034": "bond",
    "036": "mutual_fund",
    "039": "stock",
}
_INVESTMENT_WORD_TO_CLASS = {
    "obligasi": "bond",
    "reksadana": "mutual_fund",
    "saham": "stock",
}
_HARD_ASSET_HINTS = (
    "rumah",
    "apartemen",
    "ruko",
    "tanah",
    "perhiasan",
    "emas",
    "mobil",
    "motor",
    "honda",
    "toyota",
    "modal usaha",
)


def _normalize(text: str) -> str:
    s = str(text or "").strip().lower()
    s = re.sub(r"\bpt\.?\s+", "", s)
    s = s.replace("a.n.", "an").replace("a.n", "an").replace("a/n", "an")
    s = s.replace("rek.", "rek").replace("rekening", "rek")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _normalize_owner_alias(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "").replace(".", "")).strip().upper()


def _canon_owner(text: str, config: CoretaxConfig) -> str | None:
    key = _normalize_owner_alias(text)
    compact_key = key.replace(" ", "")
    for alias, canonical in config.owner_aliases.items():
        normalized_alias = _normalize_owner_alias(alias)
        if normalized_alias == key or normalized_alias.replace(" ", "") == compact_key:
            return canonical
    return None


def _canon_institution(text: str, config: CoretaxConfig) -> str | None:
    normalized = _normalize(text)
    for canonical, aliases in config.institution_aliases.items():
        for alias in aliases:
            if _normalize(alias) in normalized:
                return canonical
    return None


def _parse_keterangan(raw: str, config: CoretaxConfig, kode_harta: str = "") -> dict | None:
    normalized = _normalize(raw)
    if not normalized:
        return None

    cash_match = re.match(
        r"^tabungan\s+(?P<inst>.+?)\s+(?:kcp\s+\S+\s+)?rek\s+(?P<acct>[\w-]+)\s+an\s+(?P<owner>.+)$",
        normalized,
    )
    if cash_match:
        institution = _canon_institution(cash_match.group("inst"), config)
        owner = _canon_owner(cash_match.group("owner"), config)
        return {
            "kind": "cash",
            "institution": institution,
            "account": cash_match.group("acct"),
            "owner": owner,
        }

    if any(hint in normalized for hint in _HARD_ASSET_HINTS):
        return None

    # Asset class is determined by the kode harta first (authoritative);
    # the description word ("saham"/"obligasi"/"reksadana") is only a fallback
    # because Saham rows often write the broker name (e.g. "Indopremier
    # Sekuritas IDD ...") without ever using the word "saham".
    asset_class = _INVESTMENT_CODE_TO_CLASS.get(kode_harta)
    investment_word = next((word for word in _INVESTMENT_WORD_TO_CLASS if re.search(rf"\b{word}\b", normalized)), None)
    if asset_class is None and investment_word:
        asset_class = _INVESTMENT_WORD_TO_CLASS[investment_word]
    if asset_class is not None:
        before_owner, _, owner_text = normalized.partition(" an ")
        owner = _canon_owner(owner_text, config) if owner_text else None
        institution = _canon_institution(before_owner, config)
        idd_sid_match = re.search(r"\b(?:idd|sid)\s+([\w-]+)", normalized)
        return {
            "kind": "investment",
            "institution": institution,
            "account": None,
            "owner": owner,
            "asset_class": asset_class,
            "broker_present": institution is not None,
            "broker_token": institution,
            "idd_sid": idd_sid_match.group(1) if idd_sid_match else None,
        }

    return None


def _round_value(value: float, rounding: str) -> float:
    if rounding == "thousands":
        return round(float(value) / 1000.0) * 1000.0
    return float(value)


def _extract_template_year(path: Path, ws) -> str:
    match = re.search(r"(20\d{2})", path.stem)
    if match:
        return match.group(1)
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        for cell in row:
            cell_text = str(cell or "")
            match = re.search(r"(20\d{2})", cell_text)
            if match:
                return match.group(1)
    return "unknown"


def _validate_template(ws) -> tuple[int, int]:
    if ws is None:
        raise CoretaxTemplateError("Template workbook has no sheets")

    header_row = None
    for row_idx in range(1, 11):
        if str(ws[f"B{row_idx}"].value or "").strip() == "Kode Harta":
            header_row = row_idx
            break
    if header_row is None:
        raise CoretaxTemplateError("Template missing 'Kode Harta' header in B1:B10")

    expected_headers = {
        "D": "Tahun Perolehan",
        "G": "Nilai saat ini",
        "H": "Keterangan",
    }
    for col, header in expected_headers.items():
        cell_value = str(ws[f"{col}{header_row}"].value or "").strip().rstrip("\\")
        if cell_value != header:
            raise CoretaxTemplateError(f"Template header {header!r} must be in column {col}")
    return header_row, header_row + 1


def _load_cash_rows(conn: sqlite3.Connection, snapshot_date: str) -> list[dict]:
    rows = conn.execute(
        """
        SELECT id, institution, account, owner, currency, balance_idr
        FROM account_balances
        WHERE snapshot_date = ?
        ORDER BY institution, account, owner, id
        """,
        (snapshot_date,),
    ).fetchall()
    return [
        {
            "kind": "cash",
            "id": row["id"],
            "institution": row["institution"],
            "account": row["account"],
            "owner": row["owner"],
            "currency": row["currency"],
            "value_f": float(row["balance_idr"] or 0.0),
            "value_g": 0.0,
        }
        for row in rows
    ]


def _load_investment_rows(conn: sqlite3.Connection, snapshot_date: str) -> list[dict]:
    rows = conn.execute(
        """
        SELECT id, asset_class, institution, owner, currency, cost_basis_idr, market_value_idr
        FROM holdings
        WHERE snapshot_date = ?
        ORDER BY institution, owner, asset_class, id
        """,
        (snapshot_date,),
    ).fetchall()
    return [
        {
            "kind": "investment",
            "id": row["id"],
            "institution": row["institution"],
            "owner": row["owner"],
            "asset_class": row["asset_class"],
            "currency": row["currency"],
            "value_f": float(row["cost_basis_idr"] or 0.0),
            "value_g": float(row["market_value_idr"] or 0.0),
        }
        for row in rows
    ]


def _serialize_path(path: Path | None) -> str | None:
    return str(path) if path is not None else None


def result_to_dict(result: CoretaxResult) -> dict[str, Any]:
    data = asdict(result)
    data["output_path"] = _serialize_path(result.output_path)
    return data


def _write_audit_json(result: CoretaxResult, audit_path: Path) -> None:
    audit_path.write_text(json.dumps(result_to_dict(result), indent=2, ensure_ascii=False), encoding="utf-8")


def _next_output_path(output_dir: Path, template_year: str, snapshot_date: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = f"CoreTax_{template_year}_{snapshot_date}"
    version = 1
    while True:
        candidate = output_dir / f"{stem}_v{version}.xlsx"
        if not candidate.exists():
            return candidate
        version += 1


def _match_cash(parsed: dict, cash_rows: list[dict]) -> tuple[dict | None, list[str]]:
    warnings: list[str] = []
    if not parsed.get("institution"):
        warnings.append("could not canonicalize institution")
        return None, warnings
    if not parsed.get("owner"):
        warnings.append("could not canonicalize owner")
        return None, warnings
    # PWM has historically inconsistent owner population (some rows have owner='').
    # Strict match first; fall back to (institution, account) if the only mismatch
    # is an empty PWM owner — that's "unknown owner" which we treat as a wildcard.
    strict = [
        row for row in cash_rows
        if row["institution"] == parsed["institution"]
        and row["account"] == parsed["account"]
        and row["owner"] == parsed["owner"]
    ]
    matches = strict
    owner_wildcard = False
    if not matches:
        loose = [
            row for row in cash_rows
            if row["institution"] == parsed["institution"]
            and row["account"] == parsed["account"]
            and (row["owner"] or "") == ""
        ]
        if loose:
            matches = loose
            owner_wildcard = True
    if not matches:
        warnings.append("no matching account_balances row for institution/account/owner")
        return None, warnings
    row = matches[0]
    if owner_wildcard:
        warnings.append(f"matched on (institution, account) — PWM owner is empty; XLSX owner is {parsed['owner']!r}")
    return {
        "kind": "cash",
        "source_rows": [{
            "table": "account_balances",
            "id": row["id"],
            "institution": row["institution"],
            "account": row["account"],
            "owner": row["owner"],
            "currency": row["currency"],
            "balance_idr": row["value_f"],
        }],
        "value_f": row["value_f"],
        "value_g": 0.0,
        "currency_warning": row["currency"] != "IDR",
        "match_keys": {
            "institution": row["institution"],
            "account": row["account"],
            "owner": row["owner"],
        },
    }, warnings


def _aggregate_investments(rows: list[dict], aggregated: bool = False) -> dict:
    institutions = sorted({row["institution"] for row in rows if row.get("institution")})
    return {
        "kind": "investment",
        "source_rows": [{
            "table": "holdings",
            "id": row["id"],
            "institution": row["institution"],
            "owner": row["owner"],
            "asset_class": row["asset_class"],
            "currency": row["currency"],
            "cost_basis_idr": row["value_f"],
            "market_value_idr": row["value_g"],
        } for row in rows],
        "value_f": sum(row["value_f"] for row in rows),
        "value_g": sum(row["value_g"] for row in rows),
        "currency_warning": any(row["currency"] != "IDR" for row in rows),
        "institutions": institutions,
        "aggregated": aggregated,
    }


def _match_investment(parsed: dict, investment_rows: list[dict], config: CoretaxConfig) -> tuple[dict | None, list[str], str]:
    warnings: list[str] = []
    if not parsed.get("owner"):
        warnings.append("could not canonicalize owner")
        return None, warnings, "unmatched"
    asset_class = parsed.get("asset_class")
    owner = parsed.get("owner")
    institution = parsed.get("institution")

    if institution:
        strict_rows = [
            row for row in investment_rows
            if row["institution"] == institution
            and row["owner"] == owner
            and row["asset_class"] == asset_class
        ]
        if strict_rows:
            return _aggregate_investments(strict_rows), warnings, "filled"
        # Fallback: tolerate PWM holdings where owner is empty (data quality gap).
        owner_wildcard_rows = [
            row for row in investment_rows
            if row["institution"] == institution
            and row["asset_class"] == asset_class
            and (row["owner"] or "") == ""
        ]
        if owner_wildcard_rows:
            warnings.append(
                f"matched on (institution, asset_class) — PWM owner is empty; XLSX owner is {owner!r}"
            )
            return _aggregate_investments(owner_wildcard_rows), warnings, "filled"
        warnings.append("no matching holdings rows for institution/asset_class/owner")
        return None, warnings, "unmatched"

    if config.investment_match_mode != "aggregate_with_warning":
        warnings.append("investment row missing broker token in strict mode")
        return None, warnings, "unmatched"

    aggregate_rows = [
        row for row in investment_rows
        if row["owner"] == owner and row["asset_class"] == asset_class
    ]
    if not aggregate_rows:
        warnings.append("no matching holdings rows for asset_class/owner")
        return None, warnings, "unmatched"
    match = _aggregate_investments(aggregate_rows, aggregated=True)
    warnings.append(
        f"investment row aggregated across {len(match['institutions'])} institutions: {', '.join(match['institutions'])}"
    )
    return match, warnings, "aggregated"


def generate_coretax_xlsx(
    template_path: Path,
    output_path: Path | None,
    snapshot_date: str,
    db_path: Path,
    dry_run: bool = False,
) -> CoretaxResult:
    template_path = Path(template_path)
    db_path = Path(db_path)
    config = _CORETAX_CONFIG

    wb = openpyxl.load_workbook(template_path, keep_vba=False)
    if not wb.worksheets:
        raise CoretaxTemplateError("Template workbook has no sheets")
    ws = wb.worksheets[0]
    header_row, first_data_row = _validate_template(ws)

    conn = open_db(str(db_path))
    try:
        cash_rows = _load_cash_rows(conn, snapshot_date)
        investment_rows = _load_investment_rows(conn, snapshot_date)
    finally:
        conn.close()

    used_cash_ids: set[int] = set()
    used_investment_ids: set[int] = set()
    traces: list[CoretaxRowTrace] = []
    filled_count = 0
    skipped_out_of_scope = 0
    unmatched_count = 0
    aggregated_count = 0
    currency_warning_count = 0

    row_idx = first_data_row
    while True:
        terminator = str(ws[f"C{row_idx}"].value or "").strip()
        if terminator == "TOTAL ASET KOTOR":
            break

        kode_harta = str(ws[f"B{row_idx}"].value or "").strip()
        raw_keterangan = str(ws[f"H{row_idx}"].value or "")
        normalized = _normalize(raw_keterangan)
        parsed = _parse_keterangan(raw_keterangan, config, kode_harta=kode_harta)
        warnings: list[str] = []
        matched = None
        value_f = None
        value_g = None
        status = "unmatched"

        if kode_harta in _OUT_OF_SCOPE_CODES:
            status = "skipped_out_of_scope"
            skipped_out_of_scope += 1
        elif parsed is None:
            status = "unmatched"
            warnings.append("could not parse keterangan")
            unmatched_count += 1
        elif parsed["kind"] == "cash":
            matched, warnings = _match_cash(parsed, cash_rows)
            if matched is None:
                unmatched_count += 1
            else:
                value_f = _round_value(matched["value_f"], config.rounding)
                value_g = 0.0
                ws[f"F{row_idx}"] = value_f
                ws[f"G{row_idx}"] = value_g
                matched["value_f"] = value_f
                matched["value_g"] = value_g
                for source in matched["source_rows"]:
                    used_cash_ids.add(source["id"])
                status = "currency_warning" if matched["currency_warning"] else "filled"
                if matched["currency_warning"]:
                    currency_warning_count += 1
                    warnings.append(f"currency {matched['source_rows'][0]['currency']} converted via *_idr columns")
                filled_count += 1
        else:
            if kode_harta in _INVESTMENT_CODE_TO_CLASS and parsed.get("asset_class") != _INVESTMENT_CODE_TO_CLASS[kode_harta]:
                warnings.append(
                    f"template kode harta {kode_harta} implies {_INVESTMENT_CODE_TO_CLASS[kode_harta]} but description parsed as {parsed.get('asset_class')}"
                )
            matched, warnings, status = _match_investment(parsed, investment_rows, config)
            if matched is None:
                unmatched_count += 1
            else:
                value_f = _round_value(matched["value_f"], config.rounding)
                value_g = _round_value(matched["value_g"], config.rounding)
                ws[f"F{row_idx}"] = value_f
                ws[f"G{row_idx}"] = value_g
                matched["value_f"] = value_f
                matched["value_g"] = value_g
                for source in matched["source_rows"]:
                    used_investment_ids.add(source["id"])
                if matched.get("aggregated"):
                    status = "aggregated"
                    aggregated_count += 1
                if matched["currency_warning"]:
                    currency_warning_count += 1
                    warnings.append("currency USD converted via *_idr columns" if any(source["currency"] == "USD" for source in matched["source_rows"]) else "non-IDR currency converted via *_idr columns")
                    if status != "aggregated":
                        status = "currency_warning"
                filled_count += 1

        traces.append(
            CoretaxRowTrace(
                xlsx_row=row_idx,
                raw_keterangan=raw_keterangan,
                normalized=normalized,
                parsed=parsed,
                matched=matched,
                value_f=value_f,
                value_g=value_g,
                status=status,
                warnings=warnings,
            )
        )
        row_idx += 1

    unused_pwm_rows = [
        {
            "kind": "cash",
            "table": "account_balances",
            "id": row["id"],
            "institution": row["institution"],
            "account": row["account"],
            "owner": row["owner"],
        }
        for row in cash_rows
        if row["id"] not in used_cash_ids
    ] + [
        {
            "kind": "investment",
            "table": "holdings",
            "id": row["id"],
            "institution": row["institution"],
            "owner": row["owner"],
            "asset_class": row["asset_class"],
        }
        for row in investment_rows
        if row["id"] not in used_investment_ids
    ]

    final_output_path = None if dry_run else Path(output_path) if output_path is not None else _next_output_path(
        Path(config.output_dir),
        _extract_template_year(template_path, ws),
        snapshot_date,
    )
    if final_output_path is not None:
        final_output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(final_output_path)

    result = CoretaxResult(
        snapshot_date=snapshot_date,
        filled_count=filled_count,
        skipped_out_of_scope=skipped_out_of_scope,
        unmatched_count=unmatched_count,
        aggregated_count=aggregated_count,
        currency_warning_count=currency_warning_count,
        rows=traces,
        unused_pwm_rows=unused_pwm_rows,
        output_path=final_output_path,
    )
    if final_output_path is not None:
        _write_audit_json(result, final_output_path.with_suffix(".audit.json"))
    return result
