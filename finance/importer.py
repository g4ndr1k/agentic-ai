"""
Stage 2 — XLSX → SQLite import module.

Usage
─────
  python3 -m finance.importer              # import ALL_TRANSACTIONS.xlsx
  python3 -m finance.importer --dry-run    # parse + categorize, no writes
  python3 -m finance.importer --overwrite  # re-import; replace rows by hash
  python3 -m finance.importer --file PATH  # use a specific xlsx file
  python3 -m finance.importer -v           # verbose / debug logging

What it does
────────────
  1. Opens ALL_TRANSACTIONS.xlsx (Stage 1 output — immutable, never edited)
  2. Maps columns to the Stage 2 schema
  3. Generates a SHA-256 dedup hash per transaction
  4. Skips rows already in SQLite (hash match)
  5. Runs the 4-layer categorization engine
  6. Writes new rows directly to the SQLite transactions table
  7. Logs the import run to the import_log table

Reimport safety
───────────────
  The XLSX is the immutable ground truth.  Delete any bad rows from SQLite
  and re-run this script.  Manual category overrides survive re-import via
  the category_overrides table (resolved through the transactions_resolved view).
  Use --overwrite only when you explicitly want to replace existing rows.
"""
from __future__ import annotations
import argparse
import calendar
import difflib
import logging
import os
import re
import sys
import time
from datetime import datetime, timezone
from typing import Optional

import openpyxl

from finance.config import (
    load_config,
    get_finance_config,
    get_ollama_finance_config,
)
from finance.models import FinanceTransaction, parse_xlsx_date
from finance.categorizer import Categorizer, match_internal_transfers

log = logging.getLogger(__name__)

DEDUP_USE_ENGINE = os.environ.get("DEDUP_USE_ENGINE", "false").lower() == "true"


def _identity_key(date: str, amount: float, institution: str, account: str, owner: str) -> tuple[str, float, str, str, str]:
    return (date, round(float(amount), 2), institution, account or "", owner)


def _description_tokens(text: str) -> list[str]:
    return re.findall(r"[A-Z0-9]+", (text or "").upper())


def _ordered_prefix_token_match(shorter: list[str], longer: list[str]) -> bool:
    pos = 0
    for token in shorter:
        matched = False
        while pos < len(longer):
            current = longer[pos]
            pos += 1
            if current == token or current.startswith(token):
                matched = True
                break
        if not matched:
            return False
    return True


def _should_reconcile_parser_variant(existing_desc: str, incoming_desc: str) -> bool:
    existing_desc = (existing_desc or "").strip()
    incoming_desc = (incoming_desc or "").strip()
    if not existing_desc or not incoming_desc or existing_desc == incoming_desc:
        return False

    shorter, longer = sorted((existing_desc, incoming_desc), key=len)
    if len(longer) - len(shorter) < 12:
        return False

    shorter_tokens = _description_tokens(shorter)
    longer_tokens = _description_tokens(longer)
    if len(shorter_tokens) < 4 or len(longer_tokens) < len(shorter_tokens) + 2:
        return False

    if not _ordered_prefix_token_match(shorter_tokens, longer_tokens):
        return False

    ratio = difflib.SequenceMatcher(None, shorter, longer).ratio()
    return ratio >= 0.55


def _load_existing_transactions_by_identity(conn) -> dict[tuple[str, float, str, str, str], list[dict]]:
    rows = conn.execute(
        """
        SELECT hash, date, amount, raw_description, merchant, category,
               institution, account, owner, import_file
        FROM transactions
        """
    ).fetchall()
    out: dict[tuple[str, float, str, str, str], list[dict]] = {}
    for row in rows:
        key = _identity_key(row["date"], row["amount"], row["institution"], row["account"], row["owner"])
        out.setdefault(key, []).append(dict(row))
    return out

# ── ALL_TRANSACTIONS.xlsx column positions (0-based) ─────────────────────────
# Header row: Owner | Month | Bank | Statement Type | Tgl. Transaksi |
#             Tgl. Tercatat | Keterangan | Currency | Jumlah Valuta Asing |
#             Kurs (RP) | Jumlah (IDR) | Tipe | Saldo (IDR) | Nomor Rekening/Kartu
_C_OWNER       = 0
_C_MONTH       = 1   # not imported (derived from date)
_C_BANK        = 2
_C_STMT_TYPE   = 3   # not imported (context only)
_C_DATE_TX     = 4
_C_DATE_POST   = 5   # skipped per design (transaction date only)
_C_DESC        = 6
_C_CURRENCY    = 7
_C_FOREIGN_AMT = 8
_C_EXCHANGE_RT = 9
_C_AMOUNT_IDR  = 10
_C_TIPE        = 11
_C_BALANCE     = 12  # not imported
_C_ACCOUNT_NUM = 13

_EXPECTED_HEADERS = [
    "Owner", "Month", "Bank", "Statement Type",
    "Tgl. Transaksi", "Tgl. Tercatat", "Keterangan",
    "Currency", "Jumlah Valuta Asing", "Kurs (RP)",
    "Jumlah (IDR)", "Tipe", "Saldo (IDR)", "Nomor Rekening/Kartu",
]

# Transactions dated before this cutoff are silently dropped.
# CC billing cycles can span two calendar months, so a 2026 statement may
# contain December 2025 charges — we exclude those pre-cutoff rows here.
_MIN_TX_DATE = "2026-01-01"


def _auto_ignore_merchant(institution: str, stmt_type: str) -> str | None:
    """Return a merchant label to auto-ignore, or None if the transaction should be categorized normally."""
    inst = institution.lower()
    if inst == "ipot" and stmt_type == "statement":
        return "IPOT RDN"
    if inst == "bni sekuritas":
        return "BNIS RDN"
    if inst == "stockbit sekuritas":
        return "Stockbit RDN"
    return None


# ── Core import logic ─────────────────────────────────────────────────────────

def direct_import(
    xlsx_path: str,
    db_path: str,
    categorizer: Categorizer,
    overwrite: bool = False,
    dry_run: bool = False,
    import_file_label: str = "",
) -> dict:
    """
    Import ALL_TRANSACTIONS.xlsx directly into SQLite.

    Reads ALL_TRANSACTIONS.xlsx, categorizes each row, and writes directly
    to the ``transactions`` and ``import_log`` tables.  SQLite is the
    authoritative store.

    Dedup: INSERT OR IGNORE on the hash UNIQUE constraint (normal mode).
           UPDATE existing rows when overwrite=True.

    Returns stats dict:
        added       int   rows written to SQLite
        skipped     int   rows skipped (duplicate hash)
        total       int   total valid rows in XLSX
        parse_err   int   rows that could not be parsed
        by_layer    dict  {1: n, 2: n, 3: n, 4: n}
        duration_s  float elapsed seconds
        dry_run     bool  True if no writes were made
    """
    from finance.db import open_db

    t0 = time.time()
    label = import_file_label or os.path.basename(xlsx_path)

    # ── 1. Open XLSX ──────────────────────────────────────────────────────────
    log.info("direct_import: opening %s …", xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "ALL_TRANSACTIONS" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet 'ALL_TRANSACTIONS' not found in {xlsx_path}.")
    ws = wb["ALL_TRANSACTIONS"]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        log.warning("ALL_TRANSACTIONS sheet is empty — nothing to import.")
        return _stats(0, 0, 0, 0, {}, 0.0, dry_run)

    header = [str(v).strip() if v else "" for v in all_rows[0]]
    _check_header(header)
    data_rows = all_rows[1:]
    log.info("direct_import: read %d data rows.", len(data_rows))

    # ── 2. Load existing hashes from SQLite ───────────────────────────────────
    conn = open_db(db_path)
    existing_hashes: set[str] = {
        r[0] for r in conn.execute("SELECT hash FROM transactions").fetchall()
    }
    existing_by_identity = _load_existing_transactions_by_identity(conn)
    hash_to_category: dict[str, str] = {}
    if overwrite:
        hash_to_category = {
            r[0]: r[1]
            for r in conn.execute(
                "SELECT hash, category FROM transactions_resolved WHERE category IS NOT NULL AND category != ''"
            ).fetchall()
        }
    log.info("direct_import: %d hashes already in SQLite.", len(existing_hashes))

    # ── 3. Load aliases + categories from SQLite ──────────────────────────────
    alias_rows = conn.execute("SELECT * FROM merchant_aliases").fetchall()
    cat_rows   = conn.execute("SELECT category FROM categories").fetchall()
    conn.close()

    # ── 3b. Engine connection for dedup (Phase C) ─────────────────────────────
    engine_conn = None
    if DEDUP_USE_ENGINE and not overwrite:
        from finance.db import open_db as _open_engine_db
        from finance.matching.engine import reset_run_state
        engine_conn = _open_engine_db(db_path)
        reset_run_state()

    aliases    = [dict(r) for r in alias_rows]
    categories = [r[0] for r in cat_rows]
    categorizer.reload_aliases(aliases)
    if categories:
        categorizer.categories = categories
    log.info(
        "direct_import: %d exact aliases | %d regex aliases | %d categories",
        len(categorizer._exact), len(categorizer._regex), len(categorizer.categories),
    )

    # ── 4. Parse rows ─────────────────────────────────────────────────────────
    new_txns:       list[FinanceTransaction] = []
    overwrite_txns: list[FinanceTransaction] = []
    reconciled_txns: list[tuple[str, FinanceTransaction]] = []
    skipped      = 0
    reconciled   = 0
    parse_errors = 0
    by_layer: dict[int, int] = {1: 0, 2: 0, 3: 0, 4: 0}

    for xlsx_row_idx, raw in enumerate(data_rows, start=2):
        try:
            parse_result = _parse_row(raw, label)
        except Exception as exc:
            log.warning("Row %d: parse error — %s", xlsx_row_idx, exc)
            parse_errors += 1
            continue

        if parse_result is None:
            continue

        txn, stmt_type = parse_result
        auto_merchant = _auto_ignore_merchant(txn.institution, stmt_type)

        if txn.hash in existing_hashes:
            if overwrite:
                if auto_merchant:
                    txn.merchant = auto_merchant
                    txn.category = "Ignored"
                    by_layer[1] += 1
                else:
                    result = categorizer.categorize(txn.raw_description, owner=txn.owner, account=txn.account)
                    txn.merchant = result.merchant
                    if result.category is None and hash_to_category.get(txn.hash):
                        txn.category = hash_to_category[txn.hash]
                    else:
                        txn.category = result.category
                    by_layer[result.layer] += 1
                overwrite_txns.append(txn)
            else:
                skipped += 1
            continue

        if auto_merchant:
            txn.merchant = auto_merchant
            txn.category = "Ignored"
            by_layer[1] += 1
        else:
            result = categorizer.categorize(txn.raw_description, owner=txn.owner, account=txn.account)
            txn.merchant = result.merchant
            txn.category = result.category
            by_layer[result.layer] += 1

        identity_key = _identity_key(txn.date, txn.amount, txn.institution, txn.account, txn.owner)
        identity_matches = existing_by_identity.get(identity_key, [])
        if not overwrite and txn.hash not in existing_hashes and len(identity_matches) == 1:
            existing = identity_matches[0]
            if engine_conn is not None:
                from finance.matching.engine import classify as _engine_classify
                from finance.matching.domains.dedup import domain as _dedup_domain
                _txn_row = {
                    "date": txn.date, "amount": txn.amount,
                    "institution": txn.institution, "account": txn.account or "",
                    "owner": txn.owner, "raw_description": txn.raw_description,
                }
                _match = _engine_classify(
                    _dedup_domain, engine_conn, _txn_row, run_id=label,
                    prebuilt_indexes={"existing_by_identity": existing_by_identity},
                )
                should_reconcile = (
                    _match is not None and _match.target == f"merge:{existing['hash']}"
                )
            else:
                should_reconcile = _should_reconcile_parser_variant(
                    existing.get("raw_description", ""), txn.raw_description
                )
            if should_reconcile:
                reconciled_txns.append((existing["hash"], txn))
                existing["raw_description"] = txn.raw_description
                existing["merchant"] = txn.merchant
                existing["category"] = txn.category
                existing["import_file"] = txn.import_file
                existing_hashes.add(existing["hash"])
                reconciled += 1
                continue
        new_txns.append(txn)

    if engine_conn is not None:
        engine_conn.close()
        engine_conn = None

    total_valid = len(data_rows) - parse_errors

    all_txns = new_txns + overwrite_txns
    cross_matched = match_internal_transfers(all_txns)

    log.info(
        "direct_import: %d valid | %d new | %d skipped | %d parse errors",
        total_valid, len(new_txns), skipped, parse_errors,
    )
    log.info(
        "Categorization: L1=%d L2=%d L3=%d L4=%d cross=%d",
        by_layer[1], by_layer[2], by_layer[3], by_layer[4], cross_matched,
    )

    if dry_run:
        stats = _stats(0, skipped, total_valid, parse_errors, by_layer, time.time() - t0, dry_run=True, reconciled=reconciled)
        stats["would_add"] = len(new_txns)
        return stats

    # ── 5. Write to SQLite ────────────────────────────────────────────────────
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    added = 0

    conn = open_db(db_path)
    try:
        conn.execute("BEGIN")
        for txn in new_txns:
            cur = conn.execute(
                """INSERT OR IGNORE INTO transactions
                   (date, amount, original_currency, original_amount, exchange_rate,
                    raw_description, merchant, category, institution, account, owner,
                    notes, hash, import_date, import_file, synced_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    txn.date, txn.amount, txn.original_currency, txn.original_amount,
                    txn.exchange_rate, txn.raw_description, txn.merchant, txn.category,
                    txn.institution, txn.account, txn.owner, "",
                    txn.hash, txn.import_date, txn.import_file, now,
                ),
            )
            if cur.rowcount > 0:
                added += 1

        for existing_hash, txn in reconciled_txns:
            conn.execute(
                """
                UPDATE transactions SET
                   raw_description=?, merchant=?, category=?,
                   original_currency=?, original_amount=?, exchange_rate=?,
                   institution=?, account=?, owner=?, import_date=?, import_file=?, synced_at=?
                WHERE hash=?
                """,
                (
                    txn.raw_description, txn.merchant, txn.category,
                    txn.original_currency, txn.original_amount, txn.exchange_rate,
                    txn.institution, txn.account, txn.owner, txn.import_date, txn.import_file, now,
                    existing_hash,
                ),
            )

        if overwrite and overwrite_txns:
            for txn in overwrite_txns:
                conn.execute(
                    """UPDATE transactions SET
                       merchant=?, category=?, raw_description=?,
                       institution=?, account=?, owner=?, synced_at=?
                       WHERE hash=?""",
                    (
                        txn.merchant, txn.category, txn.raw_description,
                        txn.institution, txn.account, txn.owner, now, txn.hash,
                    ),
                )

        duration = time.time() - t0
        notes = (
            f"direct_import L1={by_layer[1]} L2={by_layer[2]} "
            f"L3={by_layer[3]} review={by_layer[4]} reconciled={reconciled}"
        )
        conn.execute(
            """INSERT INTO import_log (import_date, import_file, rows_added,
               rows_skipped, rows_total, duration_s, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (now, label, added, skipped, total_valid, round(duration, 2), notes),
        )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    sync_grogol_2_from_transactions(db_path)
    log.info("direct_import: wrote %d new rows to SQLite.", added)
    return _stats(added, skipped, total_valid, parse_errors, by_layer, time.time() - t0, reconciled=reconciled)


def sync_grogol_2_from_transactions(db_path: str) -> int:
    from finance.db import open_db

    baseline_snapshot = "2026-01-31"
    asset_class = "real_estate"
    asset_group = "Real Estate"
    asset_name = "Grogol 2"
    notes = "Auto-synced from Teguh Pranoto Chen transfer transactions."
    today = datetime.now().strftime("%Y-%m-%d")

    def month_end(date_s: str) -> str:
        dt = datetime.strptime(date_s, "%Y-%m-%d")
        return f"{dt.year:04d}-{dt.month:02d}-{calendar.monthrange(dt.year, dt.month)[1]:02d}"

    def iter_month_ends(start_s: str, end_s: str):
        start = datetime.strptime(start_s, "%Y-%m-%d")
        end = datetime.strptime(end_s, "%Y-%m-%d")
        year, month = start.year, start.month
        while (year, month) <= (end.year, end.month):
            yield f"{year:04d}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
            month += 1
            if month > 12:
                year += 1
                month = 1

    conn = open_db(db_path)
    try:
        txns = conn.execute(
            """
            SELECT date, amount
            FROM transactions
            WHERE UPPER(raw_description) LIKE '%TEGUH PRANOTO CHEN%'
              AND date >= ?
            ORDER BY date
            """,
            (baseline_snapshot,),
        ).fetchall()

        candidate_dates = {baseline_snapshot}
        for table in ("account_balances", "holdings", "liabilities", "net_worth_snapshots"):
            candidate_dates.update(
                row[0]
                for row in conn.execute(f"SELECT DISTINCT snapshot_date FROM {table}").fetchall()
                if row[0]
            )
        candidate_dates.update(month_end(row["date"]) for row in txns)
        snapshot_dates = list(iter_month_ends(baseline_snapshot, max(candidate_dates)))

        cumulative_value = 0.0
        txn_list = [(row["date"], abs(float(row["amount"]))) for row in txns]
        txn_idx = 0
        last_appraised_date = baseline_snapshot

        for snapshot_date in snapshot_dates:
            while txn_idx < len(txn_list) and txn_list[txn_idx][0] <= snapshot_date:
                last_appraised_date = txn_list[txn_idx][0]
                cumulative_value += txn_list[txn_idx][1]
                txn_idx += 1

            conn.execute(
                """
                INSERT INTO holdings (
                    snapshot_date, asset_class, asset_group, asset_name, isin_or_code,
                    institution, account, owner, currency, quantity, unit_price,
                    market_value, market_value_idr, cost_basis, cost_basis_idr,
                    unrealised_pnl_idr, exchange_rate, maturity_date, coupon_rate,
                    last_appraised_date, notes, import_date
                ) VALUES (?, ?, ?, ?, '', '', '', '', 'IDR', 1, ?, ?, ?, ?, ?, 0, 1, '', 0, ?, ?, ?)
                ON CONFLICT(snapshot_date, asset_class, asset_name, owner, institution)
                DO UPDATE SET
                    asset_group = excluded.asset_group,
                    account = excluded.account,
                    currency = excluded.currency,
                    quantity = excluded.quantity,
                    unit_price = excluded.unit_price,
                    market_value = excluded.market_value,
                    market_value_idr = excluded.market_value_idr,
                    cost_basis = excluded.cost_basis,
                    cost_basis_idr = excluded.cost_basis_idr,
                    unrealised_pnl_idr = excluded.unrealised_pnl_idr,
                    exchange_rate = excluded.exchange_rate,
                    maturity_date = excluded.maturity_date,
                    coupon_rate = excluded.coupon_rate,
                    last_appraised_date = excluded.last_appraised_date,
                    notes = excluded.notes,
                    import_date = excluded.import_date
                """,
                (
                    snapshot_date,
                    asset_class,
                    asset_group,
                    asset_name,
                    cumulative_value,
                    cumulative_value,
                    cumulative_value,
                    cumulative_value,
                    cumulative_value,
                    last_appraised_date,
                    notes,
                    today,
                ),
            )

        conn.commit()
        return len(snapshot_dates)
    finally:
        conn.close()


# ── Row parser ────────────────────────────────────────────────────────────────

def _parse_row(row: tuple, import_file: str) -> Optional[tuple["FinanceTransaction", str]]:
    """
    Convert one ALL_TRANSACTIONS.xlsx data row to a (FinanceTransaction, stmt_type) tuple.
    Returns None for blank / incomplete rows (silently skipped).

    stmt_type is the raw value from the "Statement Type" column (e.g. "savings",
    "rdn", "cc") — used by the caller to apply institution-level auto-categorization.
    """
    if not row or all(v is None or v == "" for v in row):
        return None

    def cell(idx):
        return row[idx] if idx < len(row) else None

    owner       = _str(cell(_C_OWNER))
    institution = _str(cell(_C_BANK))
    stmt_type   = _str(cell(_C_STMT_TYPE)).lower()
    raw_desc    = _str(cell(_C_DESC))
    account_num = _str(cell(_C_ACCOUNT_NUM))
    tipe        = _str(cell(_C_TIPE)).lower()        # "debit" / "credit"
    currency    = _str(cell(_C_CURRENCY)).upper() or "IDR"

    if not raw_desc or not owner or not institution:
        return None

    date_str = parse_xlsx_date(cell(_C_DATE_TX))
    if not date_str:
        return None
    if date_str < _MIN_TX_DATE:
        return None  # pre-2026 transaction — drop silently

    amount_raw = _float(cell(_C_AMOUNT_IDR))
    if amount_raw is None:
        return None
    amount = -abs(amount_raw) if tipe == "debit" else abs(amount_raw)

    if currency == "IDR":
        original_currency = None
        original_amount   = None
        exchange_rate     = None
    else:
        original_currency = currency
        original_amount   = _float(cell(_C_FOREIGN_AMT))
        exchange_rate     = _float(cell(_C_EXCHANGE_RT))
        if exchange_rate is None and original_amount:
            try:
                exchange_rate = abs(amount_raw) / abs(original_amount)
            except ZeroDivisionError:
                pass

    return FinanceTransaction(
        date=date_str,
        amount=amount,
        original_currency=original_currency,
        original_amount=original_amount,
        exchange_rate=exchange_rate,
        raw_description=raw_desc,
        merchant=None,
        category=None,
        institution=institution,
        account=account_num,
        owner=owner,
        import_file=import_file,
    ), stmt_type


# ── Small helpers ─────────────────────────────────────────────────────────────

def _str(val) -> str:
    return str(val).strip() if val is not None else ""


def _float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _stats(
    added: int, skipped: int, total: int, parse_errors: int,
    by_layer: dict, duration: float, dry_run: bool = False, reconciled: int = 0,
) -> dict:
    return {
        "added":      added,
        "skipped":    skipped,
        "reconciled": reconciled,
        "total":      total,
        "parse_err":  parse_errors,
        "by_layer":   by_layer,
        "duration_s": round(duration, 2),
        "dry_run":    dry_run,
    }


def _check_header(actual: list[str]):
    mismatches = [
        (i, exp, act)
        for i, (exp, act) in enumerate(
            zip(_EXPECTED_HEADERS, actual[:len(_EXPECTED_HEADERS)])
        )
        if exp != act
    ]
    if len(actual) != len(_EXPECTED_HEADERS):
        max_len = max(len(actual), len(_EXPECTED_HEADERS))
        for i in range(min(len(actual), len(_EXPECTED_HEADERS)), max_len):
            exp = _EXPECTED_HEADERS[i] if i < len(_EXPECTED_HEADERS) else ""
            act = actual[i] if i < len(actual) else ""
            mismatches.append((i, exp, act))

    if mismatches:
        raise ValueError(
            "XLSX header mismatch:\n"
            + "\n".join(
                f"  col {i}: expected {expected!r}, got {actual_value!r}"
                for i, expected, actual_value in mismatches
            )
        )


# ── CLI entry point ───────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Import ALL_TRANSACTIONS.xlsx into SQLite.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 -m finance.importer               # standard import (skip duplicates)
  python3 -m finance.importer --dry-run     # preview without writing
  python3 -m finance.importer --overwrite   # re-import; replace rows by hash
  python3 -m finance.importer --file /path/to/file.xlsx
        """,
    )
    parser.add_argument(
        "--overwrite", action="store_true",
        help="Replace existing rows that match by hash (default: skip them)",
    )
    parser.add_argument(
        "--file", metavar="PATH",
        help="XLSX file to import (default: finance.xlsx_input from settings.toml)",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Parse and categorize; print summary without writing",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="Enable DEBUG logging",
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s  %(levelname)-7s  %(message)s",
        datefmt="%H:%M:%S",
        stream=sys.stdout,
    )

    cfg         = load_config()
    finance_cfg = get_finance_config(cfg)
    ollama_cfg  = get_ollama_finance_config(cfg)

    xlsx_path = args.file or finance_cfg.xlsx_input
    if not os.path.exists(xlsx_path):
        log.error("XLSX file not found: %s", xlsx_path)
        sys.exit(1)

    categorizer = Categorizer(
        aliases=[],
        categories=[],
        ollama_host=ollama_cfg.host,
        ollama_model=ollama_cfg.model,
        ollama_timeout=ollama_cfg.timeout_seconds,
    )

    stats = direct_import(
        xlsx_path=xlsx_path,
        db_path=finance_cfg.sqlite_db,
        categorizer=categorizer,
        overwrite=args.overwrite,
        dry_run=args.dry_run,
        import_file_label=os.path.basename(xlsx_path),
    )

    print()
    if args.dry_run:
        print(
            f"[DRY RUN]  Would add {stats.get('would_add', 0)} rows  |  "
            f"{stats['skipped']} duplicates skipped  |  "
            f"{stats['total']} total rows in XLSX"
        )
    else:
        print(
            f"Import complete\n"
            f"  Added:      {stats['added']}\n"
            f"  Skipped:    {stats['skipped']}  (already in SQLite)\n"
            f"  Reconciled: {stats['reconciled']}  (parser-evolution duplicates merged)\n"
            f"  Total XLSX: {stats['total']}\n"
            f"  Parse err:  {stats['parse_err']}\n"
            f"  Duration:   {stats['duration_s']}s\n"
            f"\n"
            f"  Categorization breakdown:\n"
            f"    L1 auto:      {stats['by_layer'].get(1, 0)}\n"
            f"    L2 auto:      {stats['by_layer'].get(2, 0)}\n"
            f"    L3 suggested: {stats['by_layer'].get(3, 0)}\n"
            f"    L4 review:    {stats['by_layer'].get(4, 0)}"
        )
        if stats["by_layer"].get(4, 0) > 0:
            print(
                f"\n  ⚠  {stats['by_layer'][4]} transaction(s) need review in the PWA."
            )


if __name__ == "__main__":
    main()
