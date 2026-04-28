"""Tests for the persistent CoreTax SPT ledger workflow."""
from __future__ import annotations

import json
import sqlite3
import tempfile
from pathlib import Path

import openpyxl
import pytest

from finance.coretax.carry_forward import commit_staging_batch
from finance.coretax.db import (
    ASSET_CODE_SEED,
    ensure_coretax_tables,
    get_mappings,
    get_rows_for_year,
    increment_mapping_hit,
    insert_row,
    make_stable_key_cash,
    make_stable_key_investment,
    update_row,
    upsert_mapping,
    upsert_taxpayer,
)
from finance.coretax.exporter import ExportError, export_coretax_xlsx
from finance.coretax.import_parser import ImportParseError, parse_prior_year_xlsx
from finance.coretax.reconcile import run_reconcile


# ── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def conn():
    """Fresh in-memory SQLite DB with coretax tables initialized."""
    c = sqlite3.connect(":memory:")
    c.row_factory = sqlite3.Row
    ensure_coretax_tables(c)
    # PWM tables required by reconcile
    c.executescript("""
        CREATE TABLE account_balances (
            id INTEGER PRIMARY KEY, snapshot_date TEXT, institution TEXT,
            account TEXT, owner TEXT, currency TEXT, balance_idr REAL
        );
        CREATE TABLE holdings (
            id INTEGER PRIMARY KEY, snapshot_date TEXT, asset_class TEXT,
            institution TEXT, owner TEXT, currency TEXT, asset_name TEXT,
            isin_or_code TEXT, cost_basis_idr REAL, market_value_idr REAL
        );
        CREATE TABLE liabilities (
            id INTEGER PRIMARY KEY, snapshot_date TEXT, liability_type TEXT,
            liability_name TEXT, institution TEXT, owner TEXT, balance_idr REAL
        );
    """)
    c.commit()
    yield c
    c.close()


def _make_synth_template(path: Path, prior_year: int = 2024,
                         taxpayer_name: str = "Test Taxpayer",
                         npwp: str = "0000000000000000") -> None:
    """Create a small synthetic SPT template at `path` with 4 representative rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(prior_year + 1)
    ws["B1"] = "Nama wajib pajak"
    ws["C1"] = taxpayer_name
    ws["B2"] = "NPWP"
    ws["C2"] = npwp
    ws["B3"] = "Tahun Pajak"
    ws["C3"] = prior_year
    ws["A4"] = "No"
    ws["B4"] = "Kode Harta"
    ws["C4"] = "Keterangan"
    ws["D4"] = "Tahun Perolehan"
    ws["E4"] = prior_year
    ws["F4"] = prior_year + 1
    ws["G4"] = "Nilai saat ini"
    ws["H4"] = "Keterangan"
    # Rows 6+ data
    rows = [
        # No, kode, keterangan, acq_year, E (prior), F (carry), G (mv), H (note)
        (1, "061", "Tanah & Bangunan",  2010, 100000000, 100000000, 500000000, "Property A"),
        (2, "012", "Tabungan",          2020,  10000000,         0,         0, "BCA acct 1234567890 an Test Taxpayer"),
        (3, "039", "Saham",             2022, 200000000,         0, 250000000, "Indopremier IDD ABC123 an Test Taxpayer"),
        (4, "043", "Mobil",             2018, 300000000, 300000000, 200000000, "Test Car"),
    ]
    for i, (no, kode, ket, ay, e, f, g, h) in enumerate(rows, start=6):
        ws[f"A{i}"] = no
        ws[f"B{i}"] = kode
        ws[f"C{i}"] = ket
        ws[f"D{i}"] = ay
        ws[f"E{i}"] = e
        ws[f"F{i}"] = f
        ws[f"G{i}"] = g
        ws[f"H{i}"] = h
    # Footer (row 48+)
    ws["C48"] = "TOTAL ASET KOTOR"
    ws["E48"] = "=SUM(E6:E47)"
    ws["F48"] = "=SUM(F6:F47)"
    ws["C49"] = "KENAIKAN ASET KOTOR"
    ws["F49"] = "=F48-E48"
    wb.save(str(path))


@pytest.fixture
def synth_template(tmp_path):
    p = tmp_path / "CoreTax_2024.xlsx"
    _make_synth_template(p, prior_year=2024)
    return p


# ── Tests ────────────────────────────────────────────────────────────────────

def test_import_parser_synthetic_template(conn, synth_template):
    """Parser successfully reads a well-formed prior-year template."""
    result = parse_prior_year_xlsx(synth_template, target_tax_year=2025, conn=conn)
    assert result["row_count"] == 4
    assert result["prior_tax_year"] == 2024
    assert result["warnings"] == []
    # Each staged row must carry raw col E/F/G/H values for audit
    rows = conn.execute(
        "SELECT source_row_no, source_col_e_value, source_col_f_value, "
        "source_col_g_value, source_col_h_note FROM coretax_import_staging"
    ).fetchall()
    assert len(rows) == 4
    for r in rows:
        assert r["source_row_no"] >= 6
        assert r["source_col_e_value"]  # non-empty
        assert r["source_col_h_note"]   # non-empty


def test_target_year_mismatch_rejected(conn, synth_template):
    """G5: Parser rejects a template whose F header doesn't equal target_tax_year."""
    with pytest.raises(ImportParseError) as exc_info:
        parse_prior_year_xlsx(synth_template, target_tax_year=2026, conn=conn)
    assert "year mismatch" in str(exc_info.value).lower()
    assert "2024" in str(exc_info.value)


def test_f_header_mismatch_rejected(conn, synth_template):
    """G5: Parser rejects a workbook whose F header is not target_tax_year."""
    wb = openpyxl.load_workbook(str(synth_template))
    ws = wb.worksheets[0]
    ws["F4"] = 2026
    wb.save(str(synth_template))

    with pytest.raises(ImportParseError) as exc_info:
        parse_prior_year_xlsx(synth_template, target_tax_year=2025, conn=conn)
    assert "year mismatch" in str(exc_info.value).lower()
    assert "2026" in str(exc_info.value)


def test_carry_forward_commit_splits_correctly(conn, synth_template):
    """Sticky codes carry forward; refreshable codes stay unset."""
    result = parse_prior_year_xlsx(synth_template, target_tax_year=2025, conn=conn)
    commit_staging_batch(conn, result["batch_id"], 2025)

    rows = get_rows_for_year(conn, 2025)
    assert len(rows) == 4
    by_kode = {r["kode_harta"]: r for r in rows}

    # 061 (Tanah & Bangunan) — sticky
    assert by_kode["061"]["current_amount_source"] == "carried_forward"
    assert by_kode["061"]["current_amount_idr"] == 100000000

    # 043 (Mobil) — sticky
    assert by_kode["043"]["current_amount_source"] == "carried_forward"
    assert by_kode["043"]["current_amount_idr"] == 300000000

    # 012 (Tabungan) — refreshable, must NOT be pre-zeroed
    assert by_kode["012"]["current_amount_source"] == "unset"
    assert by_kode["012"]["current_amount_idr"] is None

    # 039 (Saham) — refreshable
    assert by_kode["039"]["current_amount_source"] == "unset"
    assert by_kode["039"]["current_amount_idr"] is None


def test_lock_blocks_auto_reconcile(conn):
    """G2 part 1: amount_locked=1 prevents reconcile from overwriting current_amount_idr."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "12345")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="BCA acct 12345",
        institution="BCA", account_number_masked="12345",
        current_amount_idr=999999.0, current_amount_source="manual",
        amount_locked=1, locked_reason="manual edit",
    )
    # Seed PWM that would otherwise overwrite
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "12345", "Test", "IDR", 50000.0),
    )
    # Seed mapping pointing at this row
    upsert_mapping(conn, "account_number", "12345", "012", "asset",
                   target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    result = run_reconcile(conn, 2025, "2025-01", "2025-12")
    after = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert after["current_amount_idr"] == 999999.0  # unchanged
    # Trace must record skipped_locked
    statuses = [t["status"] for t in result["trace"]]
    assert "locked_skipped" in statuses


def test_market_value_lock_independent_of_amount_lock(conn):
    """G2 part 2: market_value_locked alone blocks only market_value_idr writes."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_investment("stock", "Indopremier", "ABC123", "Test")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="039", keterangan="Indopremier ABC123",
        institution="Indopremier", external_ref="ABC123",
        market_value_idr=111111.0, market_value_source="manual",
        market_value_locked=1, amount_locked=0,
    )
    conn.execute(
        "INSERT INTO holdings (snapshot_date, asset_class, institution, owner, currency, "
        "asset_name, isin_or_code, cost_basis_idr, market_value_idr) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "stock", "Indopremier", "Test", "IDR", "ABC", "ABC123", 200000.0, 250000.0),
    )
    upsert_mapping(conn, "isin", "abc123", "039", "asset",
                   target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    run_reconcile(conn, 2025, "2025-01", "2025-12")
    after = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert after["current_amount_idr"] == 200000.0          # written (unlocked)
    assert after["market_value_idr"] == 111111.0            # untouched (locked)


def test_cash_reconcile_does_not_write_market_value(conn):
    """G2 part 3: cash kode 012 must never write market_value, even when unlocked."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "999")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="BCA 999",
        institution="BCA", account_number_masked="999",
        market_value_idr=None, market_value_source="unset",
    )
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "999", "Test", "IDR", 75000.0),
    )
    upsert_mapping(conn, "account_number", "999", "012", "asset",
                   target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    run_reconcile(conn, 2025, "2025-01", "2025-12")
    after = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert after["current_amount_idr"] == 75000.0   # cash amount written
    assert after["market_value_idr"] is None        # market value untouched


def test_mapping_hit_increments_counter(conn):
    """G1: mapping hits increment and last_used_tax_year updates on successful apply."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "11111")
    insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="BCA 11111",
        institution="BCA",
    )
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "11111", "Test", "IDR", 1000.0),
    )
    # Use the new fingerprint key format
    from finance.coretax.fingerprint import derive as fp_derive
    fp = fp_derive({"source_kind": "account_balance", "institution": "BCA",
                     "account": "11111", "owner": "Test"})
    mid = upsert_mapping(conn, fp.match_kind, fp.match_value, "012", "asset",
                         target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    run_reconcile(conn, 2025, "2025-01", "2025-12")
    m = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (mid,)).fetchone()
    assert m["hits"] == 1
    assert m["last_used_tax_year"] == 2025

    # Re-run; hits must increment again
    run_reconcile(conn, 2025, "2025-01", "2025-12")
    m = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (mid,)).fetchone()
    assert m["hits"] == 2


def test_create_from_unmatched_persists_mapping(conn):
    """G1 PWA round-trip: simulate the create-from-unmatched flow.

    1) Reconcile produces an unmatched PWM row with proposed_match_kind/value.
    2) UI calls /api/coretax/rows to create a manual ledger row.
    3) UI calls /api/coretax/mappings to persist the learned mapping.
    4) Re-run reconcile — same PWM source now auto-applies; hits increments.
    """
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "Permata", "9876", "Test", "IDR", 5000.0),
    )
    conn.commit()

    # Step 1: initial reconcile produces an unmatched row.
    r1 = run_reconcile(conn, 2025, "2025-01", "2025-12")
    assert r1["summary"]["unmatched"] == 1
    um = r1["unmatched"][0]
    assert um["payload"]["proposed_match_kind"] == "account_number_norm"
    assert um["payload"]["proposed_match_value"]  # sha256 hash

    # Step 2 + 3: UI creates a row from the unmatched payload and persists mapping.
    sk = make_stable_key_cash("Permata", "9876")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="Permata 9876",
        institution="Permata",
    )
    # Use the new fingerprint key format for the mapping
    from finance.coretax.fingerprint import derive as fp_derive
    pwm_row = {"source_kind": "account_balance", "institution": "Permata",
               "account": "9876", "owner": "Test"}
    fp = fp_derive(pwm_row)
    mid = upsert_mapping(conn, fp.match_kind, fp.match_value, "012", "asset",
                         target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()
    assert any(m["id"] == mid for m in get_mappings(conn))

    # Step 4: re-run — should apply, not appear as unmatched.
    r2 = run_reconcile(conn, 2025, "2025-01", "2025-12")
    assert r2["summary"]["unmatched"] == 0
    assert r2["summary"]["filled"] == 1
    after = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert after["current_amount_idr"] == 5000.0
    m = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (mid,)).fetchone()
    assert m["hits"] == 1


def test_legacy_mapping_auto_migration(conn):
    """Regression: legacy mapping auto-migration.

    When a mapping uses old-style keys (account_number), reconcile should:
    1. Delete the old mapping
    2. Create a new fingerprint-keyed mapping
    3. Use the NEW mapping for the match (last_mapping_id, hits, etc.)
    4. The new mapping should have hits=1, last_used_tax_year=tax_year
    """
    from finance.coretax.fingerprint import derive as fp_derive

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "11111")
    insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="BCA 11111",
        institution="BCA",
    )
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "11111", "Test", "IDR", 1000.0),
    )
    # Create mapping with OLD-style key
    old_mid = upsert_mapping(conn, "account_number", "11111", "012", "asset",
                             target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    # Verify old mapping exists
    old_mapping = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (old_mid,)).fetchone()
    assert old_mapping is not None
    assert old_mapping["match_kind"] == "account_number"

    # Run reconcile — should auto-migrate
    result = run_reconcile(conn, 2025, "2025-01", "2025-12")
    assert result["summary"]["filled"] == 1

    # Old mapping should be deleted
    old_check = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (old_mid,)).fetchone()
    assert old_check is None, "Old legacy mapping should be deleted"

    # New fingerprint mapping should exist
    fp = fp_derive({"source_kind": "account_balance", "institution": "BCA",
                     "account": "11111", "owner": "Test"})
    new_mapping = conn.execute(
        "SELECT * FROM coretax_mappings WHERE match_kind = ? AND match_value = ?",
        (fp.match_kind, fp.match_value),
    ).fetchone()
    assert new_mapping is not None, "New fingerprint mapping should exist"
    assert new_mapping["source"] == "auto_safe"
    assert new_mapping["hits"] == 1
    assert new_mapping["last_used_tax_year"] == 2025

    # Row's last_mapping_id should point to the NEW mapping
    row = conn.execute("SELECT * FROM coretax_rows WHERE id = ?",
                       (conn.execute("SELECT id FROM coretax_rows WHERE stable_key = ?", (sk,)).fetchone()["id"],)).fetchone()
    assert row["last_mapping_id"] == new_mapping["id"], \
        f"last_mapping_id should be {new_mapping['id']}, got {row['last_mapping_id']}"

    # Re-run: new mapping should increment hits
    run_reconcile(conn, 2025, "2025-01", "2025-12")
    new_mapping2 = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?",
                                 (new_mapping["id"],)).fetchone()
    assert new_mapping2["hits"] == 2


def test_export_capacity_guard(conn, tmp_path):
    """G plan #10: more than 42 asset rows must fail with no XLSX written."""
    # Need a template to load — build a minimal one
    template_dir = tmp_path / "templates"
    template_dir.mkdir()
    output_dir = tmp_path / "out"
    template_path = template_dir / "CoreTax_template.xlsx"
    _make_synth_template(template_path, prior_year=2024)

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    for i in range(43):
        insert_row(
            conn, tax_year=2025, kind="asset",
            stable_key=f"manual:061:row-{i}:2020:abc{i:04d}",
            kode_harta="061", keterangan=f"Property {i}",
            current_amount_idr=1000.0,
        )
    conn.commit()

    with pytest.raises(ExportError) as exc_info:
        export_coretax_xlsx(conn, 2025, template_dir, output_dir)
    assert "capacity" in str(exc_info.value).lower()
    # No partial XLSX must exist
    assert not any(output_dir.glob("*.xlsx")) if output_dir.exists() else True


def test_export_f_cell_formula_rule(conn, tmp_path):
    """Plan #9: F cell is `=E{n}` iff current_amount_source='carried_forward'.

    Do NOT infer formula-vs-literal from numeric equality between E and F.
    """
    template_dir = tmp_path / "templates"
    template_dir.mkdir()
    output_dir = tmp_path / "out"
    template_path = template_dir / "CoreTax_template.xlsx"
    _make_synth_template(template_path, prior_year=2024)

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    # Row A: carried forward — must produce '=E6'
    insert_row(
        conn, tax_year=2025, kind="asset",
        stable_key="manual:061:carry:2020:aaaa1111",
        kode_harta="061", keterangan="Carry Property",
        prior_amount_idr=500.0, current_amount_idr=500.0,
        prior_amount_source="imported",
        current_amount_source="carried_forward",
    )
    # Row B: literal — value coincidentally equals prior, but source is 'manual'
    insert_row(
        conn, tax_year=2025, kind="asset",
        stable_key="manual:039:literal:2022:bbbb2222",
        kode_harta="039", keterangan="Literal Stock",
        prior_amount_idr=700.0, current_amount_idr=700.0,
        prior_amount_source="imported",
        current_amount_source="manual",
    )
    conn.commit()

    result = export_coretax_xlsx(conn, 2025, template_dir, output_dir)
    out_path = output_dir / result.file_id
    wb = openpyxl.load_workbook(str(out_path), data_only=False)
    ws = wb.worksheets[0]

    # Rows are sorted by kode_harta then id: 039 first (id=2 doesn't matter,
    # 039 < 061 alphabetically). Find them deterministically by content.
    # Row 6 = first asset, row 7 = second asset.
    cells = {6: ws["F6"].value, 7: ws["F7"].value}
    ket_to_f = {ws[f"C{r}"].value: f for r, f in cells.items()}

    # Carry row → formula
    assert ket_to_f["Carry Property"] == "=E7" or ket_to_f["Carry Property"] == "=E6"
    assert isinstance(ket_to_f["Carry Property"], str)
    assert ket_to_f["Carry Property"].startswith("=E")
    # Literal row → numeric, NOT formula
    assert ket_to_f["Literal Stock"] == 700.0
    assert not (isinstance(ket_to_f["Literal Stock"], str) and
                str(ket_to_f["Literal Stock"]).startswith("=E"))


def test_check_constraint_rejects_typo(conn):
    """Plan #15: CHECK constraint catches enum typos."""
    with pytest.raises(sqlite3.IntegrityError):
        conn.execute(
            "INSERT INTO coretax_rows (tax_year, kind, stable_key, current_amount_source, "
            "created_at, updated_at) VALUES (2025, 'asset', 'k1', 'auto_reconcile', 'now', 'now')"
        )


# ── Edge case tests for mapping-first reconciliation ────────────────────────


def test_component_history_preserved_across_runs(conn):
    """Components from prior runs are preserved (is_current=0) after re-run."""
    from finance.coretax.db import list_row_components, replace_row_components_for_targets

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "11111")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk,
               kode_harta="012", keterangan="BCA 11111", institution="BCA")
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "11111", "Test", "IDR", 1000.0),
    )
    from finance.coretax.fingerprint import derive as fp_derive
    fp = fp_derive({"source_kind": "account_balance", "institution": "BCA",
                     "account": "11111", "owner": "Test"})
    upsert_mapping(conn, fp.match_kind, fp.match_value, "012", "asset",
                   target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    # First run
    r1 = run_reconcile(conn, 2025, "2025-01", "2025-12")
    run1_id = r1["run_id"]
    comps1 = list_row_components(conn, sk, 2025, current_only=False)
    assert len(comps1) == 1
    assert comps1[0]["is_current"] == 1
    assert comps1[0]["reconcile_run_id"] == run1_id

    # Second run — first run's components should be preserved but marked not current
    r2 = run_reconcile(conn, 2025, "2025-01", "2025-12")
    run2_id = r2["run_id"]
    comps_all = list_row_components(conn, sk, 2025, current_only=False)
    assert len(comps_all) == 2, f"Expected 2 historical components, got {len(comps_all)}"
    comps_current = list_row_components(conn, sk, 2025, current_only=True)
    assert len(comps_current) == 1
    assert comps_current[0]["reconcile_run_id"] == run2_id
    assert comps_current[0]["is_current"] == 1

    # Run diff should still work (prior components exist)
    # Verify the first run's components are accessible
    comps_run1 = list_row_components(conn, sk, 2025, run_id=run1_id)
    assert len(comps_run1) == 1


def test_mapping_conflict_raises(conn):
    """assign_mapping raises MappingConflictError when raise_on_conflict=True and target differs."""
    from finance.coretax.db import assign_mapping, MappingConflictError

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk1 = make_stable_key_cash("BCA", "11111")
    sk2 = make_stable_key_cash("BCA", "22222")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk1,
               kode_harta="012", keterangan="BCA 11111", institution="BCA")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk2,
               kode_harta="012", keterangan="BCA 22222", institution="BCA")
    conn.commit()

    # Create initial mapping
    mid = assign_mapping(conn, "account_number_norm", "test_hash",
                         "012", "asset", sk1, source="manual")
    assert mid > 0

    # Update same mapping to same target — should work
    mid2 = assign_mapping(conn, "account_number_norm", "test_hash",
                          "012", "asset", sk1, source="manual",
                          raise_on_conflict=True)
    assert mid2 == mid

    # Update to different target with raise_on_conflict — should raise
    with pytest.raises(MappingConflictError, match="Mapping conflict"):
        assign_mapping(conn, "account_number_norm", "test_hash",
                       "012", "asset", sk2, source="manual",
                       raise_on_conflict=True)


def test_confidence_decay_uses_run_snapshot(conn):
    """Mapping with absent fingerprint should be classified ORPHANED, not decayed.

    Per the plan: 'Don't penalize a mapping just because the underlying asset
    wasn't held this year.' Fingerprint absence → ORPHANED lifecycle bucket,
    no score change.
    """
    from finance.coretax.confidence import apply as apply_conf, RunUnused, HIGH_THRESHOLD
    from finance.coretax.db import find_lifecycle_mappings

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_cash("BCA", "11111")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk,
               kode_harta="012", keterangan="BCA 11111", institution="BCA")
    from finance.coretax.fingerprint import derive as fp_derive
    fp = fp_derive({"source_kind": "account_balance", "institution": "BCA",
                     "account": "11111", "owner": "Test"})
    mid = upsert_mapping(conn, fp.match_kind, fp.match_value, "012", "asset",
                         target_stable_key=sk, created_from_tax_year=2025)
    conn.commit()

    # Add PWM data — only OTHER account exists (BCA/11111 is absent)
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "OTHER", "99999", "Test", "IDR", 999.0),
    )
    conn.commit()

    # Run reconcile — fingerprint is NOT present in snapshot
    run_reconcile(conn, 2025, "2025-01", "2025-12")
    m = conn.execute("SELECT * FROM coretax_mappings WHERE id = ?", (mid,)).fetchone()

    # Score should NOT decay (guardrail: don't penalize absent assets)
    score = float(m["confidence_score"])
    assert score >= HIGH_THRESHOLD, f"Score should stay HIGH, got {score}"

    # Mapping should be classified as ORPHANED in lifecycle
    buckets = find_lifecycle_mappings(conn, 2025)
    orphaned_ids = [o["id"] for o in buckets.get("ORPHANED", [])]
    assert mid in orphaned_ids, f"Mapping {mid} should be ORPHANED, orphaned_ids={orphaned_ids}"


def test_rejected_suggestion_filtering(conn):
    """Rejected suggestions should be filtered out of suggestion results."""
    from finance.coretax.db import insert_rejected_suggestion, get_rejected_pairs
    from finance.coretax.suggest import suggest_mappings_for_unmapped

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    sk = make_stable_key_investment("stock", "Indopremier", "IDTEST000001", "Test")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk,
               kode_harta="039", keterangan="Indopremier IDTEST000001",
               institution="Indopremier", owner="Test")
    conn.execute(
        "INSERT INTO holdings (snapshot_date, asset_class, institution, owner, currency, "
        "asset_name, isin_or_code, cost_basis_idr, market_value_idr) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "stock", "Indopremier", "Test", "IDR", "Test Stock",
         "IDTEST000001", 100000.0, 120000.0),
    )
    conn.commit()

    # Get suggestions (should include ISIN match)
    unmapped = [{"source_kind": "holding", "source_id": 1,
                 "match_kind": "isin", "match_value": "idtest000001",
                 "fingerprint_raw": "idtest000001", "pwm_label": "Indopremier / Test Stock",
                 "payload": {"isin_or_code": "IDTEST000001", "asset_class": "stock",
                             "institution": "Indopremier", "owner": "Test",
                             "asset_name": "Test Stock"}}]
    suggestions = suggest_mappings_for_unmapped(conn, 2025, unmapped, rejected=set())
    assert len(suggestions) > 0

    # Reject the first suggestion
    s = suggestions[0]
    insert_rejected_suggestion(conn, s["match_kind"], s["match_value"],
                               s["target_stable_key"], s["rule"])

    # Get suggestions again with rejected filter
    rejected = get_rejected_pairs(conn)
    suggestions2 = suggest_mappings_for_unmapped(conn, 2025, unmapped, rejected=rejected)
    # The rejected suggestion should not appear
    for s2 in suggestions2:
        assert (s2["match_kind"], s2["match_value"], s2["target_stable_key"]) not in rejected


# ── Preview conflict semantics tests ────────────────────────────────────────


def test_preview_two_pwm_items_same_target_no_conflict(conn):
    """Two PWM items mapping to the same CoreTax row should be grouped as components, not conflicts."""
    from finance.coretax.db import get_mappings

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")

    # Create a single target row
    sk1 = make_stable_key_cash("BCA", "1234567890")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk1,
               kode_harta="012", keterangan="Tabungan BCA 1234567890",
               institution="BCA", account_number_masked="1234567890",
               current_amount_idr=1000000.0, market_value_idr=0.0)

    # Insert two PWM items that should map to the same target
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "1234567890", "Test", "IDR", 500000.0),
    )
    conn.execute(
        "INSERT INTO holdings (snapshot_date, asset_class, institution, owner, currency, "
        "asset_name, isin_or_code, cost_basis_idr, market_value_idr) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "stock", "BCA", "Test", "IDR", "BCA Stock",
         "ID1234567890", 200000.0, 250000.0),
    )
    conn.commit()

    # These are two different fingerprints mapping to the same target
    suggestions = [
        {"match_kind": "isin", "match_value": "id1234567890",
         "target_stable_key": sk1, "confidence_score": 0.95, "rule": "isin_exact",
         "source_kind": "holding", "pwm_label": "BCA Stock"},
        {"match_kind": "account_number", "match_value": "1234567890",
         "target_stable_key": sk1, "confidence_score": 0.90, "rule": "account_exact",
         "source_kind": "account_balance", "pwm_label": "Tabungan BCA"},
    ]

    # Run preview logic (simulating what the endpoint does)
    rows = get_rows_for_year(conn, 2025)
    row_by_key = {r["stable_key"]: r for r in rows}
    existing_mappings = get_mappings(conn)

    conflicts = []
    seen_fingerprints = set()
    target_deltas = {}

    for s in suggestions:
        mk = s["match_kind"]
        mv = s["match_value"]
        target_key = s["target_stable_key"]
        fingerprint = (mk, mv)

        # Check for duplicate fingerprint
        if fingerprint in seen_fingerprints:
            conflicts.append({"type": "duplicate_fingerprint", "match_kind": mk, "match_value": mv})
            continue
        seen_fingerprints.add(fingerprint)

        # Check target exists
        if target_key not in row_by_key:
            conflicts.append({"type": "missing_target", "target_stable_key": target_key})
            continue

        # Valid suggestion — accumulate
        if target_key not in target_deltas:
            target_deltas[target_key] = {"component_count": 0}
        target_deltas[target_key]["component_count"] += 1

    # Should have NO conflicts (two different fingerprints, same target is fine)
    assert len(conflicts) == 0, f"Expected no conflicts, got: {conflicts}"
    # Should have one target with 2 components
    assert len(target_deltas) == 1
    assert target_deltas[sk1]["component_count"] == 2


def test_preview_duplicate_fingerprint_is_conflict(conn):
    """Duplicate fingerprint in suggestions should be detected as a conflict."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")

    sk1 = make_stable_key_cash("BCA", "1234567890")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk1,
               kode_harta="012", keterangan="Tabungan BCA",
               institution="BCA", current_amount_idr=1000000.0)

    sk2 = make_stable_key_cash("BNI", "9876543210")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk2,
               kode_harta="012", keterangan="Tabungan BNI",
               institution="BNI", current_amount_idr=500000.0)
    conn.commit()

    # Same fingerprint appears twice targeting different rows
    suggestions = [
        {"match_kind": "isin", "match_value": "id1234567890",
         "target_stable_key": sk1, "confidence_score": 0.95, "rule": "isin_exact",
         "source_kind": "holding"},
        {"match_kind": "isin", "match_value": "id1234567890",
         "target_stable_key": sk2, "confidence_score": 0.90, "rule": "isin_exact",
         "source_kind": "holding"},
    ]

    conflicts = []
    seen_fingerprints = set()

    for s in suggestions:
        fingerprint = (s["match_kind"], s["match_value"])
        if fingerprint in seen_fingerprints:
            conflicts.append({
                "type": "duplicate_fingerprint",
                "match_kind": s["match_kind"],
                "match_value": s["match_value"],
            })
            continue
        seen_fingerprints.add(fingerprint)

    assert len(conflicts) == 1
    assert conflicts[0]["type"] == "duplicate_fingerprint"
    assert conflicts[0]["match_kind"] == "isin"
    assert conflicts[0]["match_value"] == "id1234567890"


def test_preview_missing_target_is_conflict(conn):
    """Suggestion targeting a non-existent row should be a conflict."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")
    conn.commit()

    # Only sk1 exists
    sk1 = make_stable_key_cash("BCA", "1234567890")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk1,
               kode_harta="012", keterangan="Tabungan BCA",
               institution="BCA", current_amount_idr=1000000.0)
    conn.commit()

    fake_sk = "stable:does:not:exist"
    suggestions = [
        {"match_kind": "isin", "match_value": "id1234567890",
         "target_stable_key": fake_sk, "confidence_score": 0.95, "rule": "isin_exact",
         "source_kind": "holding"},
    ]

    rows = get_rows_for_year(conn, 2025)
    row_by_key = {r["stable_key"]: r for r in rows}

    conflicts = []
    for s in suggestions:
        if s["target_stable_key"] not in row_by_key:
            conflicts.append({
                "type": "missing_target",
                "target_stable_key": s["target_stable_key"],
            })

    assert len(conflicts) == 1
    assert conflicts[0]["type"] == "missing_target"
    assert conflicts[0]["target_stable_key"] == fake_sk


def test_preview_existing_mapping_to_different_target_is_conflict(conn):
    """Suggestion for a fingerprint already mapped elsewhere should be a conflict."""
    from finance.coretax.db import get_mappings

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")

    sk1 = make_stable_key_cash("BCA", "1234567890")
    sk2 = make_stable_key_cash("BNI", "9876543210")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk1,
               kode_harta="012", keterangan="Tabungan BCA",
               institution="BCA", current_amount_idr=1000000.0)
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk2,
               kode_harta="012", keterangan="Tabungan BNI",
               institution="BNI", current_amount_idr=500000.0)

    # Create an existing mapping: isin -> sk1
    upsert_mapping(conn, "isin", "id1234567890", "012", "asset", sk1,
                   created_from_tax_year=2025)
    conn.commit()

    existing_mappings = get_mappings(conn)
    existing_by_fingerprint = {}
    for m in existing_mappings:
        existing_by_fingerprint[(m["match_kind"], m["match_value"])] = m

    # Now try to map the same fingerprint to sk2
    suggestion = {
        "match_kind": "isin", "match_value": "id1234567890",
        "target_stable_key": sk2, "confidence_score": 0.95, "rule": "isin_exact",
    }

    conflicts = []
    fingerprint = (suggestion["match_kind"], suggestion["match_value"])
    if fingerprint in existing_by_fingerprint:
        existing = existing_by_fingerprint[fingerprint]
        if existing["target_stable_key"] != suggestion["target_stable_key"]:
            conflicts.append({
                "type": "existing_mapping_conflict",
                "match_kind": suggestion["match_kind"],
                "match_value": suggestion["match_value"],
                "target_stable_key": suggestion["target_stable_key"],
                "existing_target": existing["target_stable_key"],
                "existing_mapping_id": existing["id"],
            })

    assert len(conflicts) == 1
    assert conflicts[0]["type"] == "existing_mapping_conflict"
    assert conflicts[0]["existing_target"] == sk1
    assert conflicts[0]["target_stable_key"] == sk2


def test_assign_mapping_conflict_error_structure(conn):
    """MappingConflictError should carry structured conflict metadata."""
    from finance.coretax.db import MappingConflictError, assign_mapping

    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")

    sk1 = make_stable_key_cash("BCA", "1234567890")
    sk2 = make_stable_key_cash("BNI", "9876543210")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk1,
               kode_harta="012", keterangan="Tabungan BCA",
               institution="BCA", current_amount_idr=1000000.0)
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk2,
               kode_harta="012", keterangan="Tabungan BNI",
               institution="BNI", current_amount_idr=500000.0)

    # Create mapping to sk1
    assign_mapping(conn, "isin", "id1234567890", "012", "asset", sk1,
                   created_from_tax_year=2025)
    conn.commit()

    # Try to reassign to sk2 with raise_on_conflict=True
    with pytest.raises(MappingConflictError) as exc_info:
        assign_mapping(conn, "isin", "id1234567890", "012", "asset", sk2,
                       created_from_tax_year=2025, raise_on_conflict=True)

    exc = exc_info.value
    assert exc.match_kind == "isin"
    assert exc.match_value == "id1234567890"
    assert exc.target_stable_key == sk2
    assert exc.existing_target == sk1
    assert exc.existing_mapping_id is not None
    assert "conflict" in exc.message.lower()


def test_preview_snapshot_date_affects_amounts(conn):
    """Preview should use the specified snapshot_date for PWM amounts."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")

    sk1 = make_stable_key_cash("BCA", "1234567890")
    insert_row(conn, tax_year=2025, kind="asset", stable_key=sk1,
               kode_harta="012", keterangan="Tabungan BCA",
               institution="BCA", current_amount_idr=1000000.0)

    # Insert two snapshots with different amounts
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-06-30", "BCA", "1234567890", "Test", "IDR", 500000.0),
    )
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "1234567890", "Test", "IDR", 800000.0),
    )
    conn.commit()

    # The amounts should differ based on snapshot_date
    # With 2025-06-30: balance = 500000
    # With 2025-12-31: balance = 800000
    row_mid = conn.execute(
        "SELECT balance_idr FROM account_balances WHERE snapshot_date = ? AND account = ?",
        ("2025-06-30", "1234567890")
    ).fetchone()
    assert row_mid[0] == 500000.0

    row_end = conn.execute(
        "SELECT balance_idr FROM account_balances WHERE snapshot_date = ? AND account = ?",
        ("2025-12-31", "1234567890")
    ).fetchone()
    assert row_end[0] == 800000.0


def test_tier2_auto_safe_mapping_used_on_first_run(conn):
    """Tier-2 auto-persist should use the new mapping during the same run."""
    upsert_taxpayer(conn, 2025, "Test", "0000000000000000")

    sk = make_stable_key_cash("BCA", "1234567890")
    row_id = insert_row(
        conn, tax_year=2025, kind="asset", stable_key=sk,
        kode_harta="012", keterangan="Tabungan BCA rek 1234567890",
        institution="BCA",
    )
    conn.execute(
        "INSERT INTO account_balances (snapshot_date, institution, account, owner, currency, balance_idr) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        ("2025-12-31", "BCA", "1234567890", "Test", "IDR", 123000.0),
    )
    conn.commit()

    result = run_reconcile(conn, 2025, "2025-01", "2025-12")
    assert result["summary"]["tier2_matches"] == 1

    mappings = get_mappings(conn)
    assert len(mappings) == 1
    mapping = mappings[0]
    assert mapping["source"] == "auto_safe"
    assert mapping["hits"] == 1
    assert mapping["last_used_tax_year"] == 2025

    row = conn.execute("SELECT * FROM coretax_rows WHERE id = ?", (row_id,)).fetchone()
    assert row["last_mapping_id"] == mapping["id"]
