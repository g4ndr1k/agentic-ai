from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl
import pytest

from finance.db import open_db


@pytest.fixture
def coretax_module():
    import finance.coretax_export as coretax_export
    return coretax_export


@pytest.fixture
def sample_config(coretax_module):
    return coretax_module.CoretaxConfig(
        template_dir="data/coretax/templates",
        output_dir="data/coretax/output",
        investment_match_mode="strict",
        rounding="none",
        owner_aliases={
            "Emanuel G Adrianto": "Gandrik",
            "Emanuel G. Adrianto": "Gandrik",
            "Emanuel Adrianto": "Gandrik",
            "E G Adrianto": "Gandrik",
            "Dian Pratiwi": "Helen",
        },
        institution_aliases={
            "BNI Sekuritas":      ["bni sekuritas"],
            "Stockbit Sekuritas": ["stockbit sekuritas", "stockbit"],
            "CIMB Niaga":         ["cimb niaga", "cimb"],
            "IPOT":               ["ipot", "indopremier", "indo premier"],
            "BCA":                ["bca", "bank central asia"],
            "Maybank":            ["maybank", "may bank", "bank maybank"],
            "Permata":            ["permata", "bank permata"],
            "BRI":                ["bri", "bank rakyat indonesia"],
            "DBS":                ["dbs"],
        },
    )


@pytest.fixture
def seeded_db(tmp_path: Path) -> Path:
    db_path = tmp_path / "finance.db"
    conn = open_db(str(db_path))
    conn.execute(
        """
        INSERT INTO account_balances (
            snapshot_date, institution, account, account_type, asset_group, owner,
            currency, balance, balance_idr, exchange_rate, notes, import_date
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "2026-12-31", "BCA", "2171138631", "savings", "Cash & Liquid", "Gandrik",
            "IDR", 1250000, 1250000, 1.0, "", "2026-12-31 00:00:00",
        ),
    )
    conn.execute(
        """
        INSERT INTO account_balances (
            snapshot_date, institution, account, account_type, asset_group, owner,
            currency, balance, balance_idr, exchange_rate, notes, import_date
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "2026-12-31", "DBS", "999000111", "savings", "Cash & Liquid", "Helen",
            "IDR", 500000, 500000, 1.0, "", "2026-12-31 00:00:00",
        ),
    )
    conn.execute(
        """
        INSERT INTO holdings (
            snapshot_date, asset_class, asset_group, asset_name, isin_or_code, institution,
            account, owner, currency, quantity, unit_price, market_value, market_value_idr,
            cost_basis, cost_basis_idr, unrealised_pnl_idr, exchange_rate, maturity_date,
            coupon_rate, last_appraised_date, notes, import_date
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "2026-12-31", "stock", "Investments", "BBCA", "BBCA", "Stockbit Sekuritas",
            "", "Gandrik", "IDR", 10, 9000, 100000, 100000,
            90000, 90000, 10000, 1.0, "", 0, "", "", "2026-12-31 00:00:00",
        ),
    )
    conn.execute(
        """
        INSERT INTO holdings (
            snapshot_date, asset_class, asset_group, asset_name, isin_or_code, institution,
            account, owner, currency, quantity, unit_price, market_value, market_value_idr,
            cost_basis, cost_basis_idr, unrealised_pnl_idr, exchange_rate, maturity_date,
            coupon_rate, last_appraised_date, notes, import_date
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            "2026-12-31", "stock", "Investments", "BBRI", "BBRI", "IPOT",
            "", "Gandrik", "USD", 10, 19000, 200000, 300000,
            180000, 250000, 50000, 15000.0, "", 0, "", "", "2026-12-31 00:00:00",
        ),
    )
    conn.commit()
    conn.close()
    return db_path


def _make_template(path: Path, descriptions: list[str]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "Kode Harta"
    ws["D2"] = "Tahun Perolehan"
    ws["G2"] = "Nilai saat ini"
    ws["H2"] = "Keterangan"
    row = 3
    for code, description in zip(["012", "039", "039", "061", "012"], descriptions):
        ws[f"B{row}"] = code
        ws[f"C{row}"] = f"Asset {row}"
        ws[f"F{row}"] = None
        ws[f"G{row}"] = None
        ws[f"H{row}"] = description
        row += 1
    ws[f"C{row}"] = "TOTAL ASET KOTOR"
    wb.save(path)


def _trace_by_row(result, xlsx_row: int):
    for row in result.rows:
        if row.xlsx_row == xlsx_row:
            return row
    raise AssertionError(f"row {xlsx_row} not found")


def test_coretax_normalize(coretax_module):
    value = coretax_module._normalize("  TABUNGAN PT Bank Central Asia Tbk  REK. 2171138631 a.n. Emanuel   ")
    assert value == "tabungan bank central asia tbk rek 2171138631 an emanuel"


def test_coretax_parse_tabungan(coretax_module, sample_config):
    parsed = coretax_module._parse_keterangan(
        "TABUNGAN PT Bank Central Asia Tbk KCP Sudirman REK. 2171138631 a.n. Emanuel G. Adrianto",
        sample_config,
    )
    assert parsed == {
        "kind": "cash",
        "institution": "BCA",
        "account": "2171138631",
        "owner": "Gandrik",
    }


def test_coretax_canon_institution(coretax_module, sample_config):
    assert coretax_module._canon_institution("tabungan bank central asia tbk", sample_config) == "BCA"
    assert coretax_module._canon_institution("rekening indo premier sekuritas", sample_config) == "IPOT"
    assert coretax_module._canon_institution("unknown broker", sample_config) is None


def test_coretax_owner_normalize(coretax_module, sample_config):
    assert coretax_module._canon_owner("EMANUEL G ADRIANTO", sample_config) == "Gandrik"
    assert coretax_module._canon_owner("E.G. Adrianto", sample_config) == "Gandrik"


def test_coretax_strict_vs_aggregate(coretax_module, sample_config, seeded_db, tmp_path, monkeypatch):
    template = tmp_path / "CoreTax 2025.xlsx"
    _make_template(template, [
        "Tabungan BCA rek 2171138631 an Emanuel G Adrianto",
        "Saham Stockbit an Emanuel G Adrianto",
        "Saham an Emanuel G Adrianto",
        "Rumah tinggal Jakarta",
        "Tabungan DBS rek 999000111 an Dian Pratiwi",
    ])

    monkeypatch.setattr(coretax_module, "_CORETAX_CONFIG", sample_config)
    strict = coretax_module.generate_coretax_xlsx(template, None, "2026-12-31", seeded_db, dry_run=True)
    strict_row = _trace_by_row(strict, 4)
    assert strict_row.status == "filled"
    missing_broker_row = _trace_by_row(strict, 5)
    assert missing_broker_row.status == "unmatched"

    aggregate_config = sample_config.__class__(**{**sample_config.__dict__, "investment_match_mode": "aggregate_with_warning"})
    monkeypatch.setattr(coretax_module, "_CORETAX_CONFIG", aggregate_config)
    aggregate = coretax_module.generate_coretax_xlsx(template, None, "2026-12-31", seeded_db, dry_run=True)
    aggregate_row = _trace_by_row(aggregate, 5)
    assert aggregate_row.status == "aggregated"
    assert aggregate_row.value_f == 340000
    assert aggregate_row.value_g == 400000
    assert any("aggregated across 2 institutions" in warning for warning in aggregate_row.warnings)


def test_coretax_currency_warning(coretax_module, sample_config, seeded_db, tmp_path, monkeypatch):
    template = tmp_path / "CoreTax 2025.xlsx"
    _make_template(template, [
        "Tabungan BCA rek 2171138631 an Emanuel G Adrianto",
        "Saham IPOT an Emanuel G Adrianto",
        "Saham Stockbit an Emanuel G Adrianto",
        "Rumah tinggal Jakarta",
        "Tabungan DBS rek 999000111 an Dian Pratiwi",
    ])
    monkeypatch.setattr(coretax_module, "_CORETAX_CONFIG", sample_config)

    result = coretax_module.generate_coretax_xlsx(template, None, "2026-12-31", seeded_db, dry_run=True)
    row = _trace_by_row(result, 4)
    assert row.status == "currency_warning"
    assert row.value_f == 250000
    assert row.value_g == 300000
    assert any("currency USD converted via *_idr columns" in warning for warning in row.warnings)


def test_coretax_template_validation(coretax_module, sample_config, seeded_db, tmp_path, monkeypatch):
    template = tmp_path / "broken.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"] = "Kode Harta"
    ws["D2"] = "Tahun Perolehan"
    ws["F2"] = "Nilai saat ini"
    ws["H2"] = "Keterangan"
    wb.save(template)
    monkeypatch.setattr(coretax_module, "_CORETAX_CONFIG", sample_config)

    with pytest.raises(coretax_module.CoretaxTemplateError):
        coretax_module.generate_coretax_xlsx(template, None, "2026-12-31", seeded_db, dry_run=True)


def test_coretax_unused_pwm_rows(coretax_module, sample_config, seeded_db, tmp_path, monkeypatch):
    template = tmp_path / "CoreTax 2025.xlsx"
    _make_template(template, [
        "Tabungan BCA rek 2171138631 an Emanuel G Adrianto",
        "Saham Stockbit an Emanuel G Adrianto",
        "Rumah tinggal Jakarta",
        "Modal Usaha Toko",
        "Perhiasan emas",
    ])
    monkeypatch.setattr(coretax_module, "_CORETAX_CONFIG", sample_config)

    result = coretax_module.generate_coretax_xlsx(template, None, "2026-12-31", seeded_db, dry_run=True)
    unused = {(row["kind"], row["institution"], row["owner"]) for row in result.unused_pwm_rows}
    assert ("cash", "DBS", "Helen") in unused
    assert ("investment", "IPOT", "Gandrik") in unused
